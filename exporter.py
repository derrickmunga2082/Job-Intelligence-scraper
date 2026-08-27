# ==========================================================
# JOB INTELLIGENCE SCRAPER
# PROFESSIONAL EXCEL EXPORTER
# ==========================================================


import pandas as pd

import os


from openpyxl import load_workbook

from openpyxl.styles import Font

from openpyxl.utils import get_column_letter


from database import get_jobs


from config import (

    EXPORT_FOLDER,

    EXCEL_FILE,

    CSV_FILE

)





# ==========================================================
# CREATE OUTPUT FOLDER
# ==========================================================


def create_output_folder():


    if not os.path.exists(

        EXPORT_FOLDER

    ):


        os.makedirs(

            EXPORT_FOLDER

        )







# ==========================================================
# EXPORT JOBS
# ==========================================================


def export_jobs():


    create_output_folder()



    jobs = get_jobs()



    if not jobs:


        print(

            "No jobs available for export"

        )


        return





    columns = [

        "id",

        "title",

        "company",

        "category",

        "source",

        "url",

        "location",

        "country",

        "description",

        "requirements",

        "employment_type",

        "salary",

        "published_date",

        "expiry_date",

        "status",

        "date_added",

        "last_updated"

    ]





    dataframe = pd.DataFrame(

        jobs,

        columns=columns

    )






    # Export Excel

    dataframe.to_excel(

        EXCEL_FILE,

        index=False

    )



    # Export CSV

    dataframe.to_csv(

        CSV_FILE,

        index=False,

        encoding="utf-8"

    )







    # ======================================================
    # FORMAT EXCEL
    # ======================================================


    workbook = load_workbook(

        EXCEL_FILE

    )


    sheet = workbook.active



    sheet.title = "Job Opportunities"




    # Freeze header

    sheet.freeze_panes = "A2"




    # Enable filter

    sheet.auto_filter.ref = (

        sheet.dimensions

    )





    # Header formatting

    for cell in sheet[1]:


        cell.font = Font(

            bold=True

        )





    # Column sizing

    for column in sheet.columns:


        max_length = 0


        column_letter = get_column_letter(

            column[0].column

        )



        for cell in column:


            try:


                if len(

                    str(cell.value)

                ) > max_length:


                    max_length = len(

                        str(cell.value)

                    )



            except:


                pass




        sheet.column_dimensions[

            column_letter

        ].width = min(

            max_length + 3,

            50

        )





    # Make URL clickable

    for row in sheet.iter_rows(

        min_row=2,

        min_col=6,

        max_col=6

    ):


        for cell in row:


            if cell.value:


                cell.hyperlink = cell.value

                cell.style = "Hyperlink"





    workbook.save(

        EXCEL_FILE

    )





    print(

        "Export completed"

    )


    print(

        "Excel:",

        EXCEL_FILE

    )


    print(

        "CSV:",

        CSV_FILE

    )